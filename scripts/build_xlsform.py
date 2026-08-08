"""
Builds the FEDI Baseline/Needs Assessment XLSForm from
Baseline_Needs_Assessment_Questionnaire_for_Kobo_Collect.docx.

Same construction approach and validation discipline as the eHA technical
assessment's Section 3 questionnaire (see that repo's
part2_q3_odk/scripts/build_xlsform.py): built programmatically for
reviewable, diffable history rather than hand-edited; bilingual
(English/Hausa, Hausa flagged as draft); explicit skip logic instead of
implicit assumptions; cross-question values derived rather than re-asked
wherever that removes a double-entry disagreement risk, with every such
deviation from the source document's literal field list stated plainly in
DESIGN_NOTES.md rather than silently changed.

Run: python build_xlsform.py
Writes: ../form/FEDI_Baseline_v1.xlsx
"""
from pathlib import Path
import openpyxl

HERE = Path(__file__).resolve().parent
OUT = HERE.parent / "form" / "FEDI_Baseline_v1.xlsx"

HAUSA_DRAFT_NOTE = "DRAFT - needs native Hausa speaker review before deployment"

survey_rows = []
choices_rows = []


def q(**kwargs):
    survey_rows.append(kwargs)


def choice(list_name, name, label_en, label_ha):
    choices_rows.append({"list_name": list_name, "name": name,
                          "label::English (en)": label_en,
                          "label::Hausa (ha)": label_ha})


def L(en, ha):
    return {"label::English (en)": en, "label::Hausa (ha)": ha}


def H(en, ha):
    return {"hint::English (en)": en, "hint::Hausa (ha)": f"{ha} [{HAUSA_DRAFT_NOTE}]"}


def CM(en, ha):
    return {"constraint_message::English (en)": en,
            "constraint_message::Hausa (ha)": f"{ha} [{HAUSA_DRAFT_NOTE}]"}


def yn(list_name="yes_no"):
    choice(list_name, "yes", "Yes", "Ee")
    choice(list_name, "no", "No", "A'a")


choice("yes_no", "yes", "Yes", "Ee")
choice("yes_no", "no", "No", "A'a")

# ===========================================================================
# CONSENT (gates everything else - "If No, end interview")
# ===========================================================================
choice("consent_list", "yes", "Yes", "Ee")
choice("consent_list", "no", "No", "A'a")
q(type="select_one consent_list", name="consent",
  **L("Do you agree to participate in this assessment?",
      "Kun yarda ku shiga wannan kimantawa?"),
  required="yes")

q(type="calculate", name="calc_continue", calculation="${consent}='yes'")

# ===========================================================================
# ADMIN
# ===========================================================================
q(type="begin_group", name="g_admin", **L("Interview administration", "Gudanar da hira"),
  relevant="${calc_continue}='true'")

q(type="date", name="interview_date", **L("Date of interview", "Ranar hira"), required="yes")
q(type="text", name="enumerator_name", **L("Name of enumerator", "Sunan mai tambaya"), required="yes")
q(type="text", name="community",
  **L("Name of community/location", "Sunan al'umma/wuri"),
  **H("Host community, IDP camp, ward, or settlement.",
      "Al'ummar masu masauki, sansanin 'yan gudun hijira, unguwa, ko sasanni."),
  required="yes")

choice("population_group_list", "idp", "IDP", "Yan gudun hijira na cikin gida")
choice("population_group_list", "host", "Host community", "Al'ummar masu masauki")
choice("population_group_list", "returnee", "Returnee", "Mai dawowa")
choice("population_group_list", "refugee", "Refugee/asylum seeker", "Dan gudun hijira")
choice("population_group_list", "other", "Other", "Wani")
q(type="select_one population_group_list", name="population_group",
  **L("Population group of household", "Rukunin jama'ar gida"), required="yes")

q(type="text", name="household_code",
  **L("Household unique code", "Lambar musamman ta gida"),
  **H("Use FEDI project household coding system.", "A yi amfani da tsarin lambar gidan FEDI."),
  required="yes")

q(type="end_group", name="g_admin_end")

# ===========================================================================
# DEMOGRAPHICS / DISABILITY INCLUSION
# ===========================================================================
q(type="begin_group", name="g_demographics", **L("Demographics", "Bayanan jama'a"),
  relevant="${calc_continue}='true'")

q(type="text", name="respondent_name",
  **L("Name of respondent", "Sunan wanda ake tambaya"),
  **H("Follow project confidentiality protocol - initials are sufficient if that is the project's practice.",
      "A bi ka'idar sirri ta aikin - gajerun haruffa sun isa idan haka ne al'adar aikin."),
  required="yes")

choice("respondent_sex_list", "female", "Female", "Mace")
choice("respondent_sex_list", "male", "Male", "Namiji")
q(type="select_one respondent_sex_list", name="respondent_sex",
  **L("Sex of respondent", "Jinsin wanda ake tambaya"), required="yes")

# Range floor set low (10, not 18) because household_head_type below
# explicitly includes "Child-headed" - a respondent could plausibly be a
# minor. See DESIGN_NOTES.md: the response options for age_group (next
# field) do not have a bucket below 18, which is a real gap in the source
# document between these two questions - not resolved by widening the
# bucket list on my own initiative, since that list was given explicitly.
q(type="integer", name="respondent_age",
  **L("Age of respondent", "Shekarun wanda ake tambaya"),
  constraint=". >= 10 and . <= 97", **CM("Enter an age from 10 to 97.", "Shigar da shekaru daga 10 zuwa 97."),
  required="yes")

choice("age_group_list", "18_24", "18-24", "18-24")
choice("age_group_list", "25_34", "25-34", "25-34")
choice("age_group_list", "35_49", "35-49", "35-49")
choice("age_group_list", "50_plus", "50 and above", "50 da sama")
q(type="select_one age_group_list", name="respondent_age_group",
  **L("Age category of respondent", "Rukunin shekarun wanda ake tambaya"), required="yes")

# Soft flag (not block) if the two disagree - same reasoning as the eHA
# roster/household-size flag: a hard equality constraint here would just
# pressure the enumerator into changing whichever number is more
# convenient. Only meaningful for age >= 18, given the bucket-list gap above.
q(type="calculate", name="calc_age_group_expected",
  calculation="if(${respondent_age}<25,'18_24', if(${respondent_age}<35,'25_34', "
              "if(${respondent_age}<50,'35_49','50_plus')))")
q(type="note", name="note_age_group_mismatch",
  **L("⚠ The age entered does not match the age category selected. Please check both.",
      f"⚠ Shekarun da aka shigar ba su dace da rukunin da aka zaba ba. Duba duka biyu. [{HAUSA_DRAFT_NOTE}]"),
  relevant="${respondent_age}>=18 and ${respondent_age_group}!=${calc_age_group_expected}")

choice("marital_status_list", "married", "Married", "Aure")
choice("marital_status_list", "widowed", "Widowed", "Gwauruwa/Gwauro")
choice("marital_status_list", "divorced_separated", "Divorced/separated", "Kashe aure/Rabuwa")
choice("marital_status_list", "single", "Single", "Ba a yi aure ba")
choice("marital_status_list", "prefer_not_say", "Prefer not to say", "Ban so in fada ba")
q(type="select_one marital_status_list", name="marital_status",
  **L("Marital status of respondent", "Matsayin aure na wanda ake tambaya"), required="yes")

choice("respondent_status_list", "pregnant_woman", "Pregnant woman", "Mace mai ciki")
choice("respondent_status_list", "lactating_woman", "Lactating woman", "Mace mai shayarwa")
choice("respondent_status_list", "caregiver_under2", "Caregiver of child under two", "Mai kula da yaro kasa da shekara 2")
choice("respondent_status_list", "female_headed_hh", "Female-headed household", "Gidan da mace ke shugabanci")
choice("respondent_status_list", "idp", "IDP", "Dan gudun hijira na cikin gida")
choice("respondent_status_list", "host_member", "Host community member", "Memba na al'ummar masu masauki")
choice("respondent_status_list", "other", "Other", "Wani")
q(type="select_multiple respondent_status_list", name="respondent_status",
  **L("Respondent category", "Rukunin wanda ake tambaya"), required="yes")

choice("pregnancy_status_list", "yes", "Yes", "Ee")
choice("pregnancy_status_list", "no", "No", "A'a")
choice("pregnancy_status_list", "prefer_not_say", "Prefer not to say", "Ban so in fada ba")
q(type="select_one pregnancy_status_list", name="pregnancy_status",
  **L("Is the respondent currently pregnant?", "Wanda ake tambaya na da ciki a yanzu?"),
  relevant="${respondent_sex}='female'", required="yes")

choice("lactating_status_list", "yes", "Yes", "Ee")
choice("lactating_status_list", "no", "No", "A'a")
choice("lactating_status_list", "na", "Not applicable", "Ba ya aiki")
q(type="select_one lactating_status_list", name="lactating_status",
  **L("Is the respondent currently breastfeeding/lactating?", "Wanda ake tambaya na shayarwa a yanzu?"),
  relevant="${respondent_sex}='female'", required="yes")

choice("hh_head_type_list", "female_headed", "Female-headed", "Mace ke shugabanci")
choice("hh_head_type_list", "male_headed", "Male-headed", "Namiji ke shugabanci")
choice("hh_head_type_list", "child_headed", "Child-headed", "Yaro ke shugabanci")
choice("hh_head_type_list", "elderly_headed", "Elderly-headed", "Tsoho/Tsohuwa ke shugabanci")
choice("hh_head_type_list", "other", "Other", "Wani")
q(type="select_one hh_head_type_list", name="household_head_type",
  **L("Type of household headship", "Irin shugabancin gida"), required="yes")

q(type="integer", name="household_size",
  **L("How many people live in this household?", "Mutane nawa ne ke zaune a wannan gida?"),
  constraint=". >= 1 and . <= 30", **CM("Enter a number from 1 to 30.", "Shigar da lamba daga 1 zuwa 30."),
  required="yes")
q(type="integer", name="female_household_members",
  **L("Number of females in the household", "Adadin mata a gida"),
  constraint=". >= 0 and . <= 30", **CM("Enter a number from 0 to 30.", "Shigar da lamba daga 0 zuwa 30."),
  required="yes")
q(type="integer", name="male_household_members",
  **L("Number of males in the household", "Adadin maza a gida"),
  constraint=". >= 0 and . <= 30", **CM("Enter a number from 0 to 30.", "Shigar da lamba daga 0 zuwa 30."),
  required="yes")
q(type="note", name="note_hh_sex_mismatch",
  **L("⚠ Female + male members do not add up to the total household size. Please check.",
      f"⚠ Yawan mata da maza bai kai jimlar mutanen gida ba. Duba. [{HAUSA_DRAFT_NOTE}]"),
  relevant="(${female_household_members} + ${male_household_members}) != ${household_size}")

q(type="integer", name="children_under_six_months",
  **L("How many children under 6 months live in the household?", "Yara nawa 'yan kasa da wata 6 ke zaune a gida?"),
  constraint=". >= 0 and . <= 30", **CM("Enter a number from 0 to 30.", "Shigar da lamba daga 0 zuwa 30."),
  required="yes")
q(type="integer", name="children_6_23_months",
  **L("How many children aged 6-23 months live in the household?", "Yara nawa 'yan wata 6-23 ke zaune a gida?"),
  constraint=". >= 0 and . <= 30", **CM("Enter a number from 0 to 30.", "Shigar da lamba daga 0 zuwa 30."),
  required="yes")
# children_under_two: auto-derived, not separately asked (source document
# lists it as its own "integer" question). Deviation stated in
# DESIGN_NOTES.md - it is exactly the sum of the two fields just above, so
# asking a third time invites a disagreement to clean up later rather than
# preventing one.
q(type="calculate", name="children_under_two",
  calculation="${children_under_six_months} + ${children_6_23_months}")

choice("youngest_child_sex_list", "female", "Female", "Mace")
choice("youngest_child_sex_list", "male", "Male", "Namiji")
choice("youngest_child_sex_list", "na", "Not applicable", "Ba ya aiki")
q(type="select_one youngest_child_sex_list", name="youngest_child_sex",
  **L("Sex of youngest child under two", "Jinsin karamin yaro kasa da shekara 2"),
  relevant="${children_under_two}>0", required="yes")
q(type="integer", name="youngest_child_age_months",
  **L("Age of youngest child under two, in months", "Shekarun karamin yaro kasa da shekara 2, a watanni"),
  relevant="${children_under_two}>0",
  constraint=". >= 0 and . <= 23", **CM("Enter a number from 0 to 23.", "Shigar da lamba daga 0 zuwa 23."),
  required="yes")

choice("difficulty_list", "no_difficulty", "No difficulty", "Babu wahala")
choice("difficulty_list", "some_difficulty", "Some difficulty", "Wasu wahala")
choice("difficulty_list", "a_lot_difficulty", "A lot of difficulty", "Wahala mai yawa")
choice("difficulty_list", "cannot_do", "Cannot do at all", "Ba za a iya ba ko kadan")
choice("difficulty_list", "prefer_not_say", "Prefer not to say", "Ban so in fada ba")
q(type="select_one difficulty_list", name="disability_household",
  **L("Does anyone in the household have difficulty seeing, hearing, walking, remembering, self-care, or communicating?",
      "Akwai wani a gida da ke da wahalar gani, ji, tafiya, tunawa, kula da kai, ko sadarwa?"),
  required="yes")

choice("participation_difficulty_list", "no_difficulty", "No difficulty", "Babu wahala")
choice("participation_difficulty_list", "some_difficulty", "Some difficulty", "Wasu wahala")
choice("participation_difficulty_list", "a_lot_difficulty", "A lot of difficulty", "Wahala mai yawa")
choice("participation_difficulty_list", "cannot_participate", "Cannot participate without support", "Ba za a iya shiga ba tare da taimako ba")
choice("participation_difficulty_list", "prefer_not_say", "Prefer not to say", "Ban so in fada ba")
q(type="select_one participation_difficulty_list", name="respondent_participation_difficulty",
  **L("Does the respondent personally have any difficulty participating in community activities?",
      "Wanda ake tambaya da kansa yana da wahalar shiga ayyukan al'umma?"), required="yes")

choice("participation_support_list", "closer_location", "Closer activity location", "Wurin ayyuka kusa")
choice("participation_support_list", "seating_shade", "Seating/shade", "Wurin zama/inuwa")
choice("participation_support_list", "female_facilitator", "Female facilitator", "Mai jagoranci mace")
choice("participation_support_list", "transport_support", "Transport support", "Taimakon sufuri")
choice("participation_support_list", "caregiver_support", "Caregiver support", "Taimakon mai kulawa")
choice("participation_support_list", "communication_support", "Communication support", "Taimakon sadarwa")
choice("participation_support_list", "no_support", "No support needed", "Ba a bukatar taimako")
choice("participation_support_list", "other", "Other", "Wani")
# Skip logic inferred, not stated verbatim in the source doc: asking what
# support would help only makes sense once some difficulty has been
# reported. See DESIGN_NOTES.md.
q(type="select_multiple participation_support_list", name="participation_support_needed",
  **L("What support would help the respondent participate safely?",
      "Wane irin taimako zai taimaka wa wanda ake tambaya ya shiga lafiya?"),
  relevant="${respondent_participation_difficulty}!='no_difficulty'", required="yes")

q(type="end_group", name="g_demographics_end")

# ===========================================================================
# FOOD SECURITY
# ===========================================================================
q(type="begin_group", name="g_food_security", **L("Food security", "Tsaron abinci"),
  relevant="${calc_continue}='true'")

choice("food_source_list", "own_production", "Own production", "Noman kai")
choice("food_source_list", "market", "Market purchase", "Sayan kasuwa")
choice("food_source_list", "food_assistance", "Food assistance", "Taimakon abinci")
choice("food_source_list", "borrowing", "Borrowing", "Aro")
choice("food_source_list", "gifts", "Gifts", "Kyauta")
choice("food_source_list", "other", "Other", "Wani")
q(type="select_one food_source_list", name="main_food_source",
  **L("What is your household's main source of food?", "Menene babban tushen abinci na gidanku?"), required="yes")

choice("veg_source_list", "market", "Market purchase", "Sayan kasuwa")
choice("veg_source_list", "own_production", "Own production", "Noman kai")
choice("veg_source_list", "donation", "Donation/support", "Kyauta/taimako")
choice("veg_source_list", "borrowing_gifts", "Borrowing/gifts", "Aro/kyauta")
choice("veg_source_list", "rarely_consumes", "Rarely consumes vegetables", "Ba safai ake cin kayan lambu ba")
choice("veg_source_list", "other", "Other", "Wani")
q(type="select_one veg_source_list", name="current_vegetable_source",
  **L("What is your household's current main source of vegetables?", "Menene babban tushen kayan lambu na gidanku a yanzu?"),
  required="yes")

q(type="integer", name="meals_yesterday",
  **L("How many meals did adults in the household eat yesterday?", "Manya nawa ne suka ci abinci jiya, sau nawa?"),
  constraint=". >= 0 and . <= 10", **CM("Enter a number from 0 to 10.", "Shigar da lamba daga 0 zuwa 10."),
  required="yes")

# "youngest child aged 6-23 months" - gated on the specific 6-23mo subset
# of children_under_two, not the whole under-2 group. The source document
# names this age range explicitly; the under-2 relevance alone would be
# too broad (would ask about a 2-month-old's meal count).
q(type="integer", name="child_meals_yesterday",
  **L("How many times did the youngest child aged 6-23 months eat yesterday?",
      "Sau nawa karamin yaro mai wata 6-23 ya ci abinci jiya?"),
  relevant="${children_under_two}>0 and ${youngest_child_age_months}>=6",
  constraint=". >= 0 and . <= 10", **CM("Enter a number from 0 to 10.", "Shigar da lamba daga 0 zuwa 10."),
  required="yes")

choice("food_groups_list", "cereals_tubers", "Cereals/tubers", "Hatsi/rogo")
choice("food_groups_list", "pulses", "Pulses", "Wake")
choice("food_groups_list", "vegetables", "Vegetables", "Kayan lambu")
choice("food_groups_list", "fruits", "Fruits", "'Ya'yan itace")
choice("food_groups_list", "meat_fish_eggs", "Meat/fish/eggs", "Nama/kifi/kwai")
choice("food_groups_list", "milk", "Milk products", "Kayayyakin madara")
choice("food_groups_list", "oils_fats", "Oils/fats", "Mai")
choice("food_groups_list", "sugar", "Sugar", "Sukari")
choice("food_groups_list", "none", "None", "Babu")
q(type="select_multiple food_groups_list", name="food_groups",
  **L("Which food groups did the household consume yesterday?", "Wadanne irin abinci gidan ya ci jiya?"), required="yes")

q(type="select_one yes_no", name="food_shortage_past_30_days",
  **L("In the past 30 days, did your household experience days when there was not enough food?",
      "A cikin kwanaki 30 da suka wuce, gidanku ya sha fama da rashin isasshen abinci?"), required="yes")

choice("shortage_frequency_list", "rarely", "Rarely", "Da wuya")
choice("shortage_frequency_list", "sometimes", "Sometimes", "Wani lokaci")
choice("shortage_frequency_list", "often", "Often", "Sau da yawa")
choice("shortage_frequency_list", "almost_every_week", "Almost every week", "Kusan kowane mako")
q(type="select_one shortage_frequency_list", name="food_shortage_frequency",
  **L("If yes, how often did this happen?", "Idan ee, sau nawa wannan ya faru?"),
  relevant="${food_shortage_past_30_days}='yes'", required="yes")

choice("coping_strategy_list", "reduced_meal_size", "Reduced meal size", "Rage girman abinci")
choice("coping_strategy_list", "ate_less_preferred", "Ate less preferred food", "Cin abincin da ba a fi so ba")
choice("coping_strategy_list", "borrowed", "Borrowed food/money", "Aron abinci/kudi")
choice("coping_strategy_list", "adults_ate_less", "Adults ate less for children", "Manya sun rage ci saboda yara")
choice("coping_strategy_list", "skipped_meals", "Skipped meals", "Tsallake cin abinci")
choice("coping_strategy_list", "received_support", "Received support", "An sami taimako")
choice("coping_strategy_list", "sold_assets", "Sold assets", "An sayar da kayan aiki")
choice("coping_strategy_list", "other", "Other", "Wani")
q(type="select_multiple coping_strategy_list", name="coping_strategy",
  **L("What did your household do when food was not enough?", "Me gidanku ya yi lokacin da abinci bai ishe ba?"),
  relevant="${food_shortage_past_30_days}='yes'", required="yes")

q(type="end_group", name="g_food_security_end")

# ===========================================================================
# IYCF / NUTRITION
# ===========================================================================
q(type="begin_group", name="g_nutrition", **L("Nutrition and IYCF", "Abinci mai gina jiki da ciyar da yara"),
  relevant="${calc_continue}='true'")

q(type="select_one yes_no", name="iycf_knowledge",
  **L("Have you received any information on Infant and Young Child Feeding before?",
      "Kun taba samun bayani kan ciyar da jarirai da yara kanana a baya?"), required="yes")

choice("nutrition_info_source_list", "health_facility", "Health facility", "Asibiti")
choice("nutrition_info_source_list", "chw", "Community health worker", "Ma'aikacin lafiya na al'umma")
choice("nutrition_info_source_list", "ngo_session", "NGO session", "Zaman kungiyar agaji")
choice("nutrition_info_source_list", "radio", "Radio", "Rediyo")
choice("nutrition_info_source_list", "family_friends", "Family/friends", "Iyali/abokai")
choice("nutrition_info_source_list", "religious_leader", "Religious/community leader", "Shugaban addini/al'umma")
choice("nutrition_info_source_list", "no_source", "No source", "Babu tushen")
choice("nutrition_info_source_list", "other", "Other", "Wani")
q(type="select_multiple nutrition_info_source_list", name="caregiver_nutrition_info_source",
  **L("Where do you usually receive nutrition or child feeding information?",
      "A ina kuke samun bayanan abinci mai gina jiki ko ciyar da yara galibi?"), required="yes")

choice("exclusive_bf_list", "less_6", "Less than 6 months", "Kasa da wata 6")
choice("exclusive_bf_list", "six_months", "6 months", "Wata 6")
choice("exclusive_bf_list", "more_6", "More than 6 months", "Fiye da wata 6")
choice("exclusive_bf_list", "dont_know", "Do not know", "Ban sani ba")
q(type="select_one exclusive_bf_list", name="exclusive_breastfeeding",
  **L("For how many months should a baby receive only breast milk?", "Har wa yaushe jariri ya kamata ya sha nono kadai?"),
  required="yes")

choice("complementary_feeding_list", "before_6", "Before 6 months", "Kafin wata 6")
choice("complementary_feeding_list", "at_6", "At 6 months", "Da wata 6")
choice("complementary_feeding_list", "after_6", "After 6 months", "Bayan wata 6")
choice("complementary_feeding_list", "dont_know", "Do not know", "Ban sani ba")
q(type="select_one complementary_feeding_list", name="complementary_feeding",
  **L("At what age should complementary feeding start?", "A wace shekara ya kamata a fara wasu abinci baya ga nono?"),
  required="yes")

choice("child_bf_list", "yes", "Yes", "Ee")
choice("child_bf_list", "no", "No", "A'a")
choice("child_bf_list", "na", "Not applicable", "Ba ya aiki")
q(type="select_one child_bf_list", name="child_breastfed",
  **L("Is the youngest child under two currently breastfeeding?", "Karamin yaro kasa da shekara 2 na shan nono a yanzu?"),
  relevant="${children_under_two}>0", required="yes")

choice("child_feeding_status_list", "excl_bf", "Exclusively breastfed", "Nono kadai")
choice("child_feeding_status_list", "mixed_feeding", "Mixed feeding", "Cakuda ciyarwa")
choice("child_feeding_status_list", "complementary_started", "Complementary feeding started", "An fara wasu abinci")
choice("child_feeding_status_list", "not_bf", "Not breastfeeding", "Ba ya shan nono")
choice("child_feeding_status_list", "na", "Not applicable", "Ba ya aiki")
q(type="select_one child_feeding_status_list", name="child_feeding_status",
  **L("Current feeding status of youngest child under two", "Yanayin ciyarwar karamin yaro kasa da shekara 2 a yanzu"),
  relevant="${children_under_two}>0", required="yes")

choice("diet_diversity_list", "yes", "Yes", "Ee")
choice("diet_diversity_list", "no", "No", "A'a")
choice("diet_diversity_list", "dont_know", "Do not know", "Ban sani ba")
choice("diet_diversity_list", "na", "Not applicable", "Ba ya aiki")
q(type="select_one diet_diversity_list", name="child_diet_diversity",
  **L("Did the youngest child aged 6-23 months eat foods from at least four different food groups yesterday?",
      "Karamin yaro mai wata 6-23 ya ci abinci daga akalla nau'in abinci 4 daban-daban jiya?"),
  relevant="${children_under_two}>0 and ${youngest_child_age_months}>=6", required="yes")

q(type="select_one yes_no", name="tom_brown_awareness",
  **L("Have you heard of Tom Brown before?", "Kun taba jin labarin Tom Brown a baya?"), required="yes")

choice("tom_brown_prepare_list", "yes_confident", "Yes, confidently", "Ee, da tabbaci")
choice("tom_brown_prepare_list", "yes_need_support", "Yes, but need support", "Ee, amma ina bukatar taimako")
choice("tom_brown_prepare_list", "no", "No", "A'a")
q(type="select_one tom_brown_prepare_list", name="tom_brown_prepare",
  **L("Can you prepare Tom Brown using local ingredients?", "Kuna iya shirya Tom Brown da kayan gida?"),
  relevant="${tom_brown_awareness}='yes'", required="yes")

choice("nutrition_concern_list", "limited_food", "Limited food", "Karancin abinci")
choice("nutrition_concern_list", "children_not_eating_well", "Children not eating well", "Yara ba sa ci da kyau")
choice("nutrition_concern_list", "lack_diverse_food", "Lack of diverse food", "Rashin nau'ukan abinci daban-daban")
choice("nutrition_concern_list", "poor_appetite", "Poor appetite", "Rashin sha'awar ci")
choice("nutrition_concern_list", "illness", "Illness", "Rashin lafiya")
choice("nutrition_concern_list", "lack_knowledge", "Lack of knowledge on child feeding", "Rashin sani kan ciyar da yara")
choice("nutrition_concern_list", "no_concern", "No major concern", "Babu babban damuwa")
choice("nutrition_concern_list", "other", "Other", "Wani")
q(type="select_multiple nutrition_concern_list", name="nutrition_concern",
  **L("What is the main nutrition concern in your household?", "Menene babban damuwar abinci mai gina jiki a gidanku?"),
  required="yes")

q(type="end_group", name="g_nutrition_end")

# ===========================================================================
# LIVELIHOODS / AGRICULTURE
# ===========================================================================
q(type="begin_group", name="g_livelihoods", **L("Livelihoods and agriculture", "Rayuwa da noma"),
  relevant="${calc_continue}='true'")

choice("livelihood_source_list", "farming", "Farming", "Noma")
choice("livelihood_source_list", "petty_trade", "Petty trade", "Kasuwanci karami")
choice("livelihood_source_list", "daily_labour", "Daily labour", "Aikin yini")
choice("livelihood_source_list", "livestock", "Livestock", "Kiwon dabbobi")
choice("livelihood_source_list", "salary_wage", "Salary/wage", "Albashi")
choice("livelihood_source_list", "humanitarian_assistance", "Humanitarian assistance", "Taimakon jin kai")
choice("livelihood_source_list", "remittances", "Remittances", "Kudin da ake turowa")
choice("livelihood_source_list", "no_regular", "No regular livelihood", "Babu tabbataccen abinda ake dogaro")
choice("livelihood_source_list", "other", "Other", "Wani")
q(type="select_multiple livelihood_source_list", name="livelihood_source",
  **L("What is the household's main source of livelihood?", "Menene babban abin dogaro na gidan?"), required="yes")

q(type="select_one yes_no", name="agriculture_experience",
  **L("Does anyone in the household have experience in farming or gardening?",
      "Akwai wani a gida da ke da kwarewa a noma ko lambu?"), required="yes")

q(type="end_group", name="g_livelihoods_end")

# ===========================================================================
# HOME GARDEN
# ===========================================================================
q(type="begin_group", name="g_home_garden", **L("Home garden", "Lambun gida"),
  relevant="${calc_continue}='true'")

choice("garden_space_list", "yes", "Yes", "Ee")
choice("garden_space_list", "no", "No", "A'a")
choice("garden_space_list", "shared_space", "Shared space", "Wurin hadin gwiwa")
choice("garden_space_list", "temporary_space", "Temporary space", "Wurin na wucin gadi")
choice("garden_space_list", "not_sure", "Not sure", "Ba tabbas")
q(type="select_one garden_space_list", name="garden_space",
  **L("Does your household have space for a small home garden?", "Gidanku na da wurin da za a yi karamin lambu?"),
  required="yes")

choice("water_access_list", "easy", "Easy", "Sauki")
choice("water_access_list", "moderate", "Moderate", "Matsakaici")
choice("water_access_list", "difficult", "Difficult", "Wahala")
choice("water_access_list", "very_difficult", "Very difficult", "Wahala sosai")
q(type="select_one water_access_list", name="water_access",
  **L("How easy is it for your household to access water for gardening?", "Yaya sauki ne samun ruwa don noma lambu ga gidanku?"),
  required="yes")

q(type="select_one yes_no", name="garden_experience",
  **L("Have you practiced home gardening before?", "Kun taba yin lambun gida a baya?"), required="yes")

# The two response codes treated as "has usable space" for the vegetable-
# selection block below (Yes; Shared space) is a judgement call, not
# stated in the source document, which only names "households with garden
# space" as the gating group in general terms. See DESIGN_NOTES.md.
GARDEN_HAS_SPACE = "(${garden_space}='yes' or ${garden_space}='shared_space')"

q(type="note", name="veg_selection_note",
  **L("Enumerator note: each household must select at least two and not more than three vegetable types for the home garden support. Available options are amaranthus, okra, sorrel, and lettuce only.",
      f"Bayanin mai tambaya: kowane gida dole ne ya zabi akalla kayan lambu guda 2 kuma ba fiye da 3 ba don tallafin lambun gida. Zabin da ake da su su ne alayyahu, kubewa, yakuwa, da letas kawai. [{HAUSA_DRAFT_NOTE}]"),
  relevant=GARDEN_HAS_SPACE)

choice("veg_list", "amaranthus", "Amaranthus", "Alayyahu")
choice("veg_list", "okra", "Okra", "Kubewa")
choice("veg_list", "sorrel", "Sorrel", "Yakuwa")
choice("veg_list", "lettuce", "Lettuce", "Letas")
q(type="select_multiple veg_list", name="preferred_vegetables",
  **L("From the following list, which two or three vegetables would your household prefer to cultivate?",
      "Daga jerin da ke tafe, wadanne kayan lambu guda biyu ko uku gidanku ke so a nome?"),
  relevant=GARDEN_HAS_SPACE,
  constraint="count-selected(.) >= 2 and count-selected(.) <= 3",
  **CM("Select a minimum of two and a maximum of three vegetables.",
       "Zabi akalla kayan lambu biyu kuma ba fiye da uku ba."),
  required="yes")

choice("veg_choice_reason_list", "easy_to_grow", "Easy to grow", "Sauki a nome")
choice("veg_choice_reason_list", "preferred_meals", "Preferred for household meals", "An fi so don abincin gida")
choice("veg_choice_reason_list", "can_be_sold", "Can be sold", "Za a iya sayarwa")
choice("veg_choice_reason_list", "needs_less_water", "Needs less water", "Bukatar ruwa kadan")
choice("veg_choice_reason_list", "seeds_familiar", "Seeds familiar to household", "Iri sananne ga gida")
choice("veg_choice_reason_list", "recommended_others", "Recommended by others", "Wasu sun bada shawara")
choice("veg_choice_reason_list", "other", "Other", "Wani")
q(type="select_multiple veg_choice_reason_list", name="veg_choice_reason",
  **L("Why did your household choose these vegetables?", "Me ya sa gidanku ya zabi wadannan kayan lambu?"),
  relevant=GARDEN_HAS_SPACE, required="yes")

choice("barriers_garden_list", "no_seeds", "No seeds", "Babu iri")
choice("barriers_garden_list", "no_tools", "No tools", "Babu kayan aiki")
choice("barriers_garden_list", "no_water", "No water", "Babu ruwa")
choice("barriers_garden_list", "no_land", "No land/space", "Babu fili/wuri")
choice("barriers_garden_list", "lack_knowledge", "Lack of knowledge", "Rashin sani")
choice("barriers_garden_list", "insecurity", "Insecurity", "Rashin tsaro")
choice("barriers_garden_list", "pests", "Pests", "Kwari")
choice("barriers_garden_list", "other", "Other", "Wani")
q(type="select_multiple barriers_garden_list", name="main_barriers",
  **L("What are the main barriers to having a home garden?", "Menene manyan matsalolin samun lambun gida?"),
  required="yes")

q(type="end_group", name="g_home_garden_end")

# ===========================================================================
# CLIMATE-SMART AGRICULTURE / CLIMATE SHOCKS
# ===========================================================================
q(type="begin_group", name="g_climate", **L("Climate", "Yanayi"), relevant="${calc_continue}='true'")

q(type="select_one yes_no", name="climate_smart_awareness",
  **L("Have you heard of climate-smart agriculture or climate-smart gardening practices before?",
      "Kun taba jin labarin noman da ya dace da yanayi ko lambun da ya dace da yanayi a baya?"), required="yes")

choice("climate_practices_list", "compost", "Compost/organic manure", "Taki na halitta")
choice("climate_practices_list", "mulching", "Mulching", "Rufe kasa da ganye")
choice("climate_practices_list", "water_conservation", "Water conservation", "Adana ruwa")
choice("climate_practices_list", "raised_beds", "Raised beds", "Gadaje da aka daga")
choice("climate_practices_list", "mixed_cropping", "Mixed cropping", "Hada iri daban-daban")
choice("climate_practices_list", "pest_control_local", "Pest control using local methods", "Sarrafa kwari da hanyoyin gida")
choice("climate_practices_list", "drought_tolerant", "Drought-tolerant crops", "Amfanin gona mai jure fari")
choice("climate_practices_list", "none", "None", "Babu")
choice("climate_practices_list", "other", "Other", "Wani")
q(type="select_multiple climate_practices_list", name="climate_smart_practices_used",
  **L("Which climate-smart gardening practices have you used or heard about?",
      "Wadanne dabarun lambu masu dacewa da yanayi kuka yi amfani da su ko ku ji labarinsu?"), required="yes")

choice("climate_shocks_list", "windstorm", "Windstorm", "Guguwa")
choice("climate_shocks_list", "flooding", "Flooding", "Ambaliyar ruwa")
choice("climate_shocks_list", "drought", "Drought/poor rainfall", "Fari/karancin ruwan sama")
choice("climate_shocks_list", "extreme_heat", "Extreme heat", "Zafi mai tsanani")
choice("climate_shocks_list", "pest_infestation", "Pest infestation", "Yawan kwari")
choice("climate_shocks_list", "soil_degradation", "Soil degradation", "Lalacewar kasa")
choice("climate_shocks_list", "none", "None", "Babu")
choice("climate_shocks_list", "other", "Other", "Wani")
q(type="select_multiple climate_shocks_list", name="climate_shocks",
  **L("Which climate or environmental shocks have affected your household in the past 12 months?",
      "Wadanne matsalolin yanayi ko muhalli suka shafi gidanku a cikin watanni 12 da suka wuce?"), required="yes")

choice("windstorm_impact_list", "yes", "Yes", "Ee")
choice("windstorm_impact_list", "no", "No", "A'a")
choice("windstorm_impact_list", "prefer_not_say", "Prefer not to say", "Ban so in fada ba")
q(type="select_one windstorm_impact_list", name="windstorm_impact",
  **L("Was your household affected by the recent powerful windstorm in Bama?",
      "Guguwa mai karfi da ta faru a Bama kwanan nan ta shafi gidanku?"), required="yes")

choice("windstorm_damage_list", "house_damaged", "House damaged", "Gida ya lalace")
choice("windstorm_damage_list", "food_lost", "Food items lost", "An rasa kayan abinci")
choice("windstorm_damage_list", "items_lost", "Household items lost", "An rasa kayan gida")
choice("windstorm_damage_list", "livelihood_assets_damaged", "Livelihood assets damaged", "Kayan aikin rayuwa sun lalace")
choice("windstorm_damage_list", "garden_farm_affected", "Garden/farm affected", "Lambu/gona ya shafa")
choice("windstorm_damage_list", "injury", "Injury", "Rauni")
choice("windstorm_damage_list", "displacement", "Displacement", "Rasa muhalli")
choice("windstorm_damage_list", "other", "Other", "Wani")
q(type="select_multiple windstorm_damage_list", name="windstorm_damage_type",
  **L("If affected, what type of damage did your household experience?", "Idan ya shafa, wane irin lahani gidanku ya sha?"),
  relevant="${windstorm_impact}='yes'", required="yes")

q(type="end_group", name="g_climate_end")

# ===========================================================================
# PROTECTION MAINSTREAMING
# ===========================================================================
q(type="begin_group", name="g_protection", **L("Protection", "Kariya"), relevant="${calc_continue}='true'")

choice("protection_concern_list", "insecurity_movement", "Insecurity/movement restriction", "Rashin tsaro/hana motsi")
choice("protection_concern_list", "distance", "Distance to activity location", "Nisa zuwa wurin ayyuka")
choice("protection_concern_list", "care_responsibilities", "Care responsibilities", "Nauyin kulawa")
choice("protection_concern_list", "domestic_workload", "Domestic workload", "Aikin gida")
choice("protection_concern_list", "disability_barrier", "Disability/accessibility barrier", "Matsalar nakasa/samun dama")
choice("protection_concern_list", "fear_harassment", "Fear of harassment", "Tsoron cin zarafi")
choice("protection_concern_list", "no_concern", "No concern", "Babu damuwa")
choice("protection_concern_list", "other", "Other", "Wani")
q(type="select_multiple protection_concern_list", name="protection_concern",
  **L("Are there any protection or safety concerns that may affect participation in project activities?",
      "Akwai wata damuwar kariya ko tsaro da za ta iya shafar shiga ayyukan aikin?"), required="yes")

choice("safe_time_list", "morning", "Morning", "Safiya")
choice("safe_time_list", "afternoon", "Afternoon", "Rana")
choice("safe_time_list", "evening", "Evening", "Yamma")
choice("safe_time_list", "depends_day", "Depends on day", "Ya danganta da rana")
choice("safe_time_list", "not_sure", "Not sure", "Ba tabbas")
q(type="select_one safe_time_list", name="safe_participation_time",
  **L("What time is safest and most convenient for caregivers to attend sessions?",
      "Wane lokaci ne mafi aminci kuma dacewa ga masu kulawa su halarci zaman?"), required="yes")

choice("comfort_speaking_list", "yes", "Yes", "Ee")
choice("comfort_speaking_list", "no", "No", "A'a")
choice("comfort_speaking_list", "somewhat", "Somewhat", "Wani kadan")
choice("comfort_speaking_list", "not_sure", "Not sure", "Ba tabbas")
q(type="select_one comfort_speaking_list", name="women_comfort_speaking",
  **L("Would women/caregivers feel comfortable speaking during group sessions?",
      "Mata/masu kulawa za su ji dadin magana a lokacin zaman kungiya?"), required="yes")

choice("facilitator_pref_list", "female", "Female facilitator", "Mai jagoranci mace")
choice("facilitator_pref_list", "male", "Male facilitator", "Mai jagoranci namiji")
choice("facilitator_pref_list", "either", "Either is fine", "Ko dai daya ya dace")
choice("facilitator_pref_list", "not_sure", "Not sure", "Ba tabbas")
q(type="select_one facilitator_pref_list", name="preferred_facilitator",
  **L("For sensitive discussions, do participants prefer a female or male facilitator?",
      "Don tattaunawa mai laushi, mahalarta sun fi son mai jagoranci mace ko namiji?"), required="yes")

choice("feedback_channel_comfort_list", "yes", "Yes", "Ee")
choice("feedback_channel_comfort_list", "no", "No", "A'a")
choice("feedback_channel_comfort_list", "need_explanation", "Need explanation", "Ana bukatar bayani")
q(type="select_one feedback_channel_comfort_list", name="feedback_channel",
  **L("Would you be comfortable using a phone number or community focal point to share feedback or complaints?",
      "Za ku ji dadin amfani da lambar waya ko wakilin al'umma don bayar da ra'ayi ko korafi?"), required="yes")

choice("feedback_pref_list", "phone_call", "Phone call", "Kiran waya")
choice("feedback_pref_list", "sms_whatsapp", "SMS/WhatsApp", "SMS/WhatsApp")
choice("feedback_pref_list", "female_focal_point", "Female community focal point", "Wakilin al'umma mace")
choice("feedback_pref_list", "male_focal_point", "Male community focal point", "Wakilin al'umma namiji")
choice("feedback_pref_list", "suggestion_box", "Suggestion box", "Akwatin shawarwari")
choice("feedback_pref_list", "during_meeting", "During meeting", "A lokacin taro")
choice("feedback_pref_list", "prefer_not_say", "Prefer not to say", "Ban so in fada ba")
q(type="select_multiple feedback_pref_list", name="feedback_preference",
  **L("What is the safest and most comfortable way for your household to share feedback or complaints?",
      "Menene hanya mafi aminci kuma dacewa ga gidanku don bayar da ra'ayi ko korafi?"), required="yes")

q(type="end_group", name="g_protection_end")

# ===========================================================================
# CHILD EDUCATION (esp. girls)
# ===========================================================================
q(type="begin_group", name="g_education", **L("Child education", "Ilimin yara"), relevant="${calc_continue}='true'")

q(type="integer", name="children_school_age_total",
  **L("How many school-age children are in the household?", "Yara nawa masu shekarun makaranta ke gida?"),
  constraint=". >= 0 and . <= 30", **CM("Enter a number from 0 to 30.", "Shigar da lamba daga 0 zuwa 30."),
  required="yes")
q(type="integer", name="school_age_girls_total",
  **L("How many school-age girls are in the household?", "'Yan mata nawa masu shekarun makaranta ke gida?"),
  relevant="${children_school_age_total}>0",
  constraint=". >= 0 and . <= 30", **CM("Enter a number from 0 to 30.", "Shigar da lamba daga 0 zuwa 30."),
  required="yes")
q(type="integer", name="school_age_boys_total",
  **L("How many school-age boys are in the household?", "'Yan yara maza nawa masu shekarun makaranta ke gida?"),
  relevant="${children_school_age_total}>0",
  constraint=". >= 0 and . <= 30", **CM("Enter a number from 0 to 30.", "Shigar da lamba daga 0 zuwa 30."),
  required="yes")
q(type="note", name="note_school_age_mismatch",
  **L("⚠ Girls + boys of school age do not add up to the total school-age children. Please check.",
      f"⚠ Yawan 'yan mata da maza masu shekarun makaranta bai kai jimlar ba. Duba. [{HAUSA_DRAFT_NOTE}]"),
  relevant="${children_school_age_total}>0 and "
           "(${school_age_girls_total} + ${school_age_boys_total}) != ${children_school_age_total}")

q(type="integer", name="girls_attending_school",
  **L("How many school-age girls are currently attending school or learning activities?",
      "'Yan mata nawa masu shekarun makaranta ke halartar makaranta ko ayyukan koyo a yanzu?"),
  relevant="${school_age_girls_total}>0",
  constraint=". >= 0 and . <= ${school_age_girls_total}",
  **CM("Cannot exceed the number of school-age girls recorded above.",
       "Ba za ta wuce yawan 'yan mata masu shekarun makaranta da aka rubuta a sama ba."),
  required="yes")
q(type="integer", name="boys_attending_school",
  **L("How many school-age boys are currently attending school or learning activities?",
      "Yara maza nawa masu shekarun makaranta ke halartar makaranta ko ayyukan koyo a yanzu?"),
  relevant="${school_age_boys_total}>0",
  constraint=". >= 0 and . <= ${school_age_boys_total}",
  **CM("Cannot exceed the number of school-age boys recorded above.",
       "Ba za ta wuce yawan yara maza masu shekarun makaranta da aka rubuta a sama ba."),
  required="yes")

# girls_out_of_school and girls_school_attendance: auto-derived rather than
# separately asked (source doc lists both as their own fields - integer and
# select_one respectively). Deviation stated in DESIGN_NOTES.md: both are
# exact functions of school_age_girls_total and girls_attending_school, so
# asking them separately only creates two more chances for the categorical
# answer to disagree with the counted one.
q(type="calculate", name="girls_out_of_school",
  calculation="${school_age_girls_total} - ${girls_attending_school}")
q(type="calculate", name="girls_school_attendance",
  calculation="if(${school_age_girls_total}=0,'no_girls', "
              "if(${girls_attending_school}=${school_age_girls_total},'all_attending', "
              "if(${girls_attending_school}=0,'none_attending','some_attending')))")

choice("girls_barriers_list", "cost_materials", "Cost of school materials", "Kudin kayan makaranta")
choice("girls_barriers_list", "household_chores", "Household chores", "Aikin gida")
choice("girls_barriers_list", "early_marriage_risk", "Early marriage risk", "Hadarin auren wuri")
choice("girls_barriers_list", "safety_concerns", "Safety concerns", "Damuwar tsaro")
choice("girls_barriers_list", "distance_school", "Distance to school", "Nisan makaranta")
choice("girls_barriers_list", "lack_female_teachers", "Lack of female teachers", "Rashin malamai mata")
choice("girls_barriers_list", "displacement", "Displacement", "Rasa muhalli")
choice("girls_barriers_list", "disability_barrier", "Disability/accessibility barrier", "Matsalar nakasa/samun dama")
choice("girls_barriers_list", "no_barrier", "No barrier", "Babu matsala")
choice("girls_barriers_list", "other", "Other", "Wani")
q(type="select_multiple girls_barriers_list", name="girls_education_barriers",
  **L("What are the main barriers affecting girls' education in the household or community?",
      "Menene manyan matsalolin da ke shafar ilimin 'yan mata a gida ko al'umma?"),
  relevant="${school_age_girls_total}>0", required="yes")

q(type="end_group", name="g_education_end")

# ===========================================================================
# CLOSE-OUT (incl. auto-derived vulnerability summary)
# ===========================================================================
q(type="begin_group", name="g_closeout", **L("Close-out", "Kammalawa"), relevant="${calc_continue}='true'")

# vulnerability_summary: auto-derived from answers already captured, not a
# manually-completed select_multiple as the source document lists it.
# Deviation stated in DESIGN_NOTES.md - this is the same clerical-recount
# risk class fixed repeatedly in the eHA build (a human re-deriving, from
# memory, a summary of facts already on the form). The computed string is
# shown back to the enumerator (note field below) rather than hidden, so
# it can still be sanity-checked in the field.
q(type="calculate", name="vulnerability_summary",
  calculation=(
      "concat("
      "if(selected(${respondent_status},'female_headed_hh') or ${household_head_type}='female_headed','female_headed_hh ','') , "
      "if(${pregnancy_status}='yes' or ${lactating_status}='yes','pregnant_lactating ',''), "
      "if(${children_under_two}>0,'caregiver_under2 ',''), "
      "if(${disability_household}!='no_difficulty','disability_in_hh ',''), "
      "if(${windstorm_impact}='yes','windstorm_affected ',''), "
      "if(selected(${livelihood_source},'no_regular'),'no_regular_livelihood ',''), "
      "if(${school_age_girls_total}>0 and ${girls_attending_school}<${school_age_girls_total},'girls_out_of_school ',''), "
      "if(${food_shortage_past_30_days}='yes','food_shortage ',''), "
      "if(${garden_space}='no' or ${garden_space}='temporary_space','limited_garden_space ',''), "
      "if(not(selected(${protection_concern},'no_concern')),'protection_concern ','')"
      ")"
  ))
q(type="note", name="note_vulnerability_summary_display",
  **L("Computed vulnerability flags for this household (for review, not editable): ${vulnerability_summary}",
      f"Alamomin rauni da aka lissafa don wannan gida (don dubawa, ba za a iya gyarawa ba): "
      f"${{vulnerability_summary}} [{HAUSA_DRAFT_NOTE}]"))

choice("photo_consent_list", "yes", "Yes", "Ee")
choice("photo_consent_list", "no", "No", "A'a")
choice("photo_consent_list", "only_no_face", "Only without showing face", "Kawai ba tare da nuna fuska ba")
q(type="select_one photo_consent_list", name="photo_consent",
  **L("Do you consent to photos being taken during project activities for reporting/visibility purposes?",
      "Kun yarda a dauki hotuna a lokacin ayyukan aikin don rahoto/nuna aiki?"), required="yes")

# Not required, per the source document's own hint ("Collect only if safe
# and appropriate") - forcing entry would directly contradict that hint.
q(type="geopoint", name="gps_location",
  **L("GPS location of household/community point", "Wurin GPS na gida/wurin al'umma"),
  **H("Collect only if safe and appropriate.", "A dauka kawai idan yana da aminci kuma ya dace."))

q(type="text", name="observations",
  **L("Enumerator observations", "Lura na mai tambaya"),
  **H("Note vulnerability, garden space, safety issues, protection considerations, and any relevant follow-up needs.",
      "A rubuta rauni, wurin lambu, matsalolin tsaro, la'akari da kariya, da duk wani bukatar bibiya."),
  appearance="multiline")

q(type="end_group", name="g_closeout_end")

# ===========================================================================


def main():
    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    ws_survey = wb.active
    ws_survey.title = "survey"
    ws_choices = wb.create_sheet("choices")
    ws_settings = wb.create_sheet("settings")

    survey_cols = ["type", "name", "label::English (en)", "label::Hausa (ha)",
                   "hint::English (en)", "hint::Hausa (ha)", "required",
                   "constraint", "constraint_message::English (en)",
                   "constraint_message::Hausa (ha)", "relevant", "calculation",
                   "appearance"]
    ws_survey.append(survey_cols)
    for row in survey_rows:
        ws_survey.append([row.get(c, "") for c in survey_cols])

    choices_cols = ["list_name", "name", "label::English (en)", "label::Hausa (ha)"]
    ws_choices.append(choices_cols)
    for row in choices_rows:
        ws_choices.append([row.get(c, "") for c in choices_cols])

    ws_settings.append(["form_title", "form_id", "version", "default_language", "style"])
    ws_settings.append(["FEDI Baseline/Needs Assessment", "fedi_baseline_v1",
                         "2026080800", "Hausa (ha)", "pages"])

    wb.save(OUT)
    print(f"Wrote {OUT} — {len(survey_rows)} survey rows, {len(choices_rows)} choice rows")


if __name__ == "__main__":
    main()
